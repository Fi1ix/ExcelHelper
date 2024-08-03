import openpyxl
from tkinter import Tk, filedialog, messagebox

def import_excel():
    root = Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx; *.xls")])
    if not file_path:
        messagebox.showwarning("警告", "未选择任何文件")
        return None
    else:
        messagebox.showinfo("文件选择", f"已选择文件: {file_path}")
        return file_path

def process_excel(input_file):
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active

    # 获取每行数据，并找出最大列数
    max_cols = 0
    data = []
    for row in sheet.iter_rows(values_only=True):
        max_cols = max(max_cols, len(row))
        data.append(row)

    # 将每行数据扩展为相同长度
    for i in range(len(data)):
        while len(data[i]) < max_cols:
            data[i] += ('',)

    # 创建新的工作簿并写入数据
    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active
    for row in data:
        new_sheet.append(row)

    # 保存新的Excel文件
    output_file = 'output.xlsx'
    new_wb.save(output_file)
    messagebox.showinfo("操作完成", f"已生成新的Excel文件: {output_file}")

if __name__ == "__main__":
    input_file = import_excel()
    if input_file:
        process_excel(input_file)
